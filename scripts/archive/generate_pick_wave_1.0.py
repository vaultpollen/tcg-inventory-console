import csv
import sqlite3
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

import sys
from pathlib import Path
sys.path.append(str(Path(__file__).resolve().parents[1]))
from core.paths import (
    DB_PATH,
    OUT_DIR,
    PICK_WAVE_CSV,
    PACK_ORDER_CSV,
    MISSING_ITEMS_CSV,
    PICK_WORKFLOW_CSV,
    PICK_WAVE_XLSX,
)
from core.location import resolve_batch_location

def norm_name(s: str) -> str:
    s = (s or "").lower().strip()
    if " - " in s:
        s = s.split(" - ")[0]
    for ch in ["-", "/", ".", ",", ":", "(", ")"]:
        s = s.replace(ch, " ")
    return " ".join(s.split())


def fuzzy_name_ok(pull_name: str, inv_name: str) -> bool:
    a = norm_name(pull_name)
    b = norm_name(inv_name)

    if not a or not b:
        return False
    if a == b:
        return True
    if a in b or b in a:
        return True
    return set(a.split()).issubset(set(b.split()))


def fetch_pull_items(conn):
    return conn.execute("""
        SELECT id, order_id, quantity, card_name, collector_no, print, condition, tcg, set_name, rarity
        FROM pull_items
        ORDER BY id
    """).fetchall()


def fetch_inventory_candidates(conn, item):
    return conn.execute("""
        SELECT
            c.row_id,
            c.name,
            c.collector_no,
            c.print,
            c.condition,
            c.language,
            c.quantity,
            c.batch_id,
            c.set_code,
            c.set_name,
            c.rarity,
            c.tcg
        FROM cards_raw c
        WHERE
            c.quantity > 0
            AND c.collector_no = ?
            AND lower(c.print) = lower(?)
            AND lower(c.condition) = lower(?)
            AND lower(c.tcg) = lower(?)
        ORDER BY
            c.batch_id,
            c.quantity DESC
    """, (
        item["collector_no"],
        item["print"],
        item["condition"],
        item["tcg"],
    )).fetchall()


def allocate(conn):
    pull_items = fetch_pull_items(conn)

    pick_rows = []
    pack_rows = []
    missing_rows = []
    workflow_rows = []

    for item in pull_items:
        need = int(item["quantity"])
        remaining = need

        candidates = fetch_inventory_candidates(conn, item)
        matched = [c for c in candidates if fuzzy_name_ok(item["card_name"], c["name"])]

        for c in matched:
            if remaining <= 0:
                break

            available = int(c["quantity"] or 0)
            if available <= 0:
                continue

            pick_qty = min(available, remaining)
            remaining -= pick_qty

            loc = resolve_batch_location(conn, c["batch_id"])

            box_id = loc["box_id"] if loc else ""
            segment = loc["segment"] if loc else ""

            location = f"{box_id}:{segment}" if (box_id or segment) else ""

            row = {
                "pull_id": item["id"],
                "order_id": item["order_id"],
                "step": item["id"],
                "source": "DB PICK",
                "card_name": item["card_name"],
                "inventory_name": c["name"],
                "collector_no": item["collector_no"],
                "print": item["print"],
                "condition": item["condition"],
                "qty_needed": need,
                "qty_to_pick": pick_qty,
                "row_id": c["row_id"],
                "batch_id": c["batch_id"],
                "box_id": box_id,
                "segment": segment,
                "location": location,
                "physical_location": loc["physical_location"] if loc else "",
                "location_source": loc["source_table"] if loc else "",
                "set_name": c["set_name"] or "",
                "status": "OK",
            }

            pick_rows.append(row)
            pack_rows.append(row.copy())
            workflow_rows.append(row.copy())

        if remaining > 0:
            fallback = {
                "pull_id": item["id"],
                "order_id": item["order_id"],
                "step": item["id"],
                "source": "TCGA / LEGACY",
                "card_name": item["card_name"],
                "inventory_name": "",
                "collector_no": item["collector_no"],
                "print": item["print"],
                "condition": item["condition"],
                "qty_needed": need,
                "qty_to_pick": remaining,
                "row_id": "",
                "batch_id": "",
                "box_id": "",
                "segment": "",
                "location": "",
                "set_name": item["set_name"] or "",
                "status": "NOT IN DB - CHECK TCGA / LEGACY",
            }
            
            workflow_rows.append(fallback.copy())
            pack_rows.append(fallback.copy())

            missing_rows.append({
                "pull_id": item["id"],
                "order_id": item["order_id"],
                "card_name": item["card_name"],
                "collector_no": item["collector_no"],
                "print": item["print"],
                "condition": item["condition"],
                "qty_needed": need,
                "qty_allocated": need - remaining,
                "qty_missing": remaining,
                "set_name": item["set_name"] or "",
                "status": f"MISSING {remaining}",
            })

    return pick_rows, pack_rows, missing_rows, workflow_rows


def write_csv(path, rows, fieldnames):
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def autosize(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            max_len = max(max_len, len(str(cell.value or "")))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)


def style_header(ws):
    fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def write_table(ws, rows, fields):
    ws.append(fields)
    for r in rows:
        ws.append([r.get(f, "") for f in fields])
    style_header(ws)
    autosize(ws)


def write_workflow_sheet(ws, rows):
    fields = [
        "order_id", "step", "source", "qty_to_pick",
        "card_name", "inventory_name", "collector_no",
        "print", "condition", "batch_id", "location",
        "row_id", "set_name", "status"
    ]

    ws.append(fields)
    style_header(ws)

    fill_db = PatternFill("solid", fgColor="E2F0D9")
    fill_tcga = PatternFill("solid", fgColor="FFF2CC")
    fill_missing = PatternFill("solid", fgColor="FFC7CE")

    current_order = None

    for r in rows:
        if r["order_id"] != current_order:
            current_order = r["order_id"]
            ws.append([])
            ws.append([f"ORDER {current_order}"])
            row_num = ws.max_row
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=len(fields))
            ws.cell(row_num, 1).font = Font(bold=True, size=12)
            ws.cell(row_num, 1).fill = PatternFill("solid", fgColor="B7DEE8")

        ws.append([r.get(f, "") for f in fields])
        row_num = ws.max_row

        if r["source"] == "DB PICK":
            fill = fill_db
        elif r["source"] == "TCGA CHECK":
            fill = fill_tcga
        else:
            fill = fill_missing

        for c in range(1, len(fields) + 1):
            ws.cell(row_num, c).fill = fill

    autosize(ws)


def write_pick_wave_sheet(ws, rows):
    fields = [
        "box_id", "segment", "location", "batch_id",
        "qty_to_pick", "card_name", "inventory_name",
        "collector_no", "print", "condition", "order_id",
        "row_id", "set_name", "status"
    ]

    ws.append(fields)
    style_header(ws)

    current_group = None
    fills = [
        PatternFill("solid", fgColor="FFFFFF"),
        PatternFill("solid", fgColor="F2F2F2"),
    ]
    group_idx = 0

    for r in rows:
        group = (r["box_id"], r["segment"], r["batch_id"])

        if group != current_group:
            current_group = group
            group_idx += 1

            ws.append([])
            header = f"BOX {r['box_id']} | SEGMENT {r['segment']} | BATCH {r['batch_id']} | LOCATION {r['location']}"
            ws.append([header])
            row_num = ws.max_row
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=len(fields))
            cell = ws.cell(row_num, 1)
            cell.font = Font(bold=True, size=12)
            cell.fill = PatternFill("solid", fgColor="B7DEE8")

        ws.append([r.get(f, "") for f in fields])
        row_num = ws.max_row
        fill = fills[group_idx % 2]
        for c in range(1, len(fields) + 1):
            ws.cell(row_num, c).fill = fill

    autosize(ws)


def write_xlsx(pick_rows, pack_rows, missing_rows, workflow_rows):
    wb = Workbook()

    ws_workflow = wb.active
    ws_workflow.title = "Pick Workflow"
    write_workflow_sheet(ws_workflow, workflow_rows)

    #ws_pick = wb.create_sheet("Pick Wave")
    #write_pick_wave_sheet(ws_pick, pick_rows)

    ws_pack = wb.create_sheet("Pack Order")
    pack_fields = [
        "order_id", "source", "qty_to_pick", "card_name", "inventory_name",
        "collector_no", "print", "condition", "batch_id",
        "location", "row_id", "status"
    ]
    write_table(ws_pack, pack_rows, pack_fields)

    ws_missing = wb.create_sheet("Missing")
    missing_fields = [
        "order_id", "card_name", "collector_no", "print", "condition",
        "qty_needed", "qty_allocated", "qty_missing", "set_name", "status"
    ]
    write_table(ws_missing, missing_rows, missing_fields)

    red_fill = PatternFill("solid", fgColor="FFC7CE")
    for row in ws_missing.iter_rows(min_row=2):
        for cell in row:
            cell.fill = red_fill

    wb.save(PICK_WAVE_XLSX)


def main():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row

    try:
        pick_rows, pack_rows, missing_rows, workflow_rows = allocate(conn)
    finally:
        conn.close()

    pick_rows.sort(key=lambda r: (
        str(r["box_id"]),
        str(r["segment"]),
        str(r["batch_id"]),
        str(r["order_id"]),
        str(r["card_name"]),
    ))

    pack_rows.sort(key=lambda r: (
        str(r["order_id"]),
        int(r["pull_id"]),
        str(r["card_name"]),
    ))

    workflow_rows.sort(key=lambda r: (
        0 if r["source"] == "DB PICK" else 1,
        str(r["box_id"]),
        str(r["segment"]),
        str(r["batch_id"]),
        str(r["order_id"]),
        int(r["pull_id"]),
        str(r["card_name"]),
    ))

    pick_fields = [
        "pull_id",
        "box_id", "segment", "location", "batch_id",
        "qty_to_pick", "card_name", "inventory_name",
        "collector_no", "print", "condition", "order_id",
        "row_id", "set_name", "status"
    ]

    pack_fields = [
        "order_id", "qty_to_pick", "card_name", "inventory_name",
        "collector_no", "print", "condition", "batch_id",
        "location", "row_id", "status"
    ]

    missing_fields = [
        "order_id", "card_name", "collector_no", "print", "condition",
        "qty_needed", "qty_allocated", "qty_missing", "set_name", "status"
    ]

    workflow_fields = [
        "order_id", "step", "source", "qty_to_pick",
        "card_name", "inventory_name", "collector_no",
        "print", "condition", "batch_id", "location",
        "row_id", "set_name", "status"
    ]

    write_csv(PICK_WAVE_CSV, pick_rows, pick_fields)
    write_csv(PACK_ORDER_CSV, pack_rows, pack_fields)
    write_csv(MISSING_ITEMS_CSV, missing_rows, missing_fields)
    write_csv(PICK_WORKFLOW_CSV, workflow_rows, workflow_fields)

    write_xlsx(pick_rows, pack_rows, missing_rows, workflow_rows)

    print(f"pick_wave rows: {len(pick_rows)}")
    print(f"pack_order rows: {len(pack_rows)}")
    print(f"workflow rows: {len(workflow_rows)}")
    print(f"missing rows: {len(missing_rows)}")
    print(f"Excel output: {PICK_WAVE_XLSX.resolve()}")


if __name__ == "__main__":
    main()
